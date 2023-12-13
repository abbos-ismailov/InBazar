from aiogram import types
from aiogram.types import Message
from data.config import ADMINS
from loader import dp, db, bot

from data.prodcut_shablon import REGULAR_SHIPPING, FAST_SHIPPING, PICKUP_SHIPPING, clothes_pr

from openpyxl import load_workbook

from docxtpl import DocxTemplate
from datetime import datetime as dt

hour = dt.now().strftime("%H")
minutes = dt.now().strftime("%M")
day = dt.now().strftime("%d")
month = dt.now().strftime("%m")
year = dt.now().strftime("%Y")

path = './inbazar_data/files/clothes_data.xlsx'
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active


# Userga Invoice taklif qildik
@dp.message_handler(text="To'lov qilish")
async def show_invoices(msg: Message):
    cl_class = await clothes_pr(msg.from_user.id)
    try:
        await bot.send_invoice(chat_id=msg.from_user.id, **cl_class.generate_invoice(), payload="Kiyimlar")
    except:
        print("Send invoice da xatolik")


# Yetkazib berish xizmati
@dp.shipping_query_handler()
async def choose_shipping(query: types.ShippingQuery):
    if query.shipping_address.country_code != "UZ":
        await bot.answer_shipping_query(shipping_query_id=query.id,
                                        ok=False,
                                        error_message="Chet elga yetkazib bera olmaymiz...")

    elif query.shipping_address.city.lower() == 'toshkent' or query.shipping_address.city.lower() == 'tashkent':
        await bot.answer_shipping_query(shipping_query_id=query.id,
                                        shipping_options=[
                                            FAST_SHIPPING, REGULAR_SHIPPING, PICKUP_SHIPPING],
                                        ok=True
                                        )
    else:
        await bot.answer_shipping_query(shipping_query_id=query.id,
                                        shipping_options=[REGULAR_SHIPPING],
                                        ok=True
                                        )


# Xarid qilingandan keyingi function
@dp.pre_checkout_query_handler()
async def process_pre_checkout_query(checkout_query: types.PreCheckoutQuery):
    await bot.answer_pre_checkout_query(pre_checkout_query_id=checkout_query.id, ok=True)
    
    await bot.send_message(chat_id=checkout_query.from_user.id, text="Xaridingiz uchun rahmat")

    info_products = await db.show_purchases(checkout_query.from_user.id) 
    info_text = ""
    
    for i in info_products:
        info_text += f"<b>{i['product_name'].title()}</b> <b>{i['product_count']}</b> ta ID: {i['product_id']}\n"    

    
    
    desc = f"Quyidagi mahsulot sotildi:\n{info_text}\nID: {checkout_query.id}\n"
    desc += f"Xaridor: {checkout_query.order_info.name}, tel: {checkout_query.order_info.phone_number}\n"
    desc += f"Telegram User: <a href='https://t.me/{checkout_query.from_user.username}'>{checkout_query.from_user.full_name}</a>"
    await bot.send_message(chat_id=ADMINS[0], text=desc)
    
    
    ### <------------ Bu foydalanuvchiga docx chiqarib beradi -------------------->
    ### Chiqqanda ochmayapti
    # data = DocxTemplate('inbazar_data/files/check.docx')
    
    # obj = {
    #     'name': info_text
    # }
    # data.render(obj)
    # data.save(f"inbazar_data/files/{checkout_query.from_user.full_name}-{day}-{month}-{year}.docx")
    
    
    id = sheet_obj.max_row + 1
    index = 0
    for i in range(id, len(info_products) + id):
        sheet_obj[f"A{i}"].value = info_products[index]["product_id"]
        sheet_obj[f"B{i}"].value = info_products[index]["product_count"]
        sheet_obj[f"C{i}"].value = info_products[index]["product_name"]
        sheet_obj[f"D{i}"].value = checkout_query.from_user.id
        sheet_obj[f"E{i}"].value = checkout_query.from_user.username
        index += 1
    
    wb_obj.save("./inbazar_data/files/clothes_data.xlsx")
    
    for i in info_products:
        await db.delete_product_from_korzinka(i['product_id'], i['user_tg_id'])